import os
import json
from datetime import datetime, timedelta
from io import BytesIO
from flask import current_app
import csv
import zipfile

# Try to import reportlab for PDF generation
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

# Try to import openpyxl for Excel generation
try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Try to import xlsxwriter for additional Excel features
try:
    import xlsxwriter
    XLSXWRITER_AVAILABLE = True
except ImportError:
    XLSXWRITER_AVAILABLE = False

# Try to import other optional dependencies
try:
    import markdown
    MARKDOWN_AVAILABLE = True
except ImportError:
    MARKDOWN_AVAILABLE = False

try:
    from jinja2 import Template
    JINJA2_AVAILABLE = True
except ImportError:
    JINJA2_AVAILABLE = False

from models import Indicator, db
from sqlalchemy import func, and_, or_

class ReportGenerator:
    def __init__(self):
        self.reports_dir = os.path.join(current_app.root_path, 'static', 'reports')
        os.makedirs(self.reports_dir, exist_ok=True)
        self.styles = getSampleStyleSheet()
        self.setup_custom_styles()
    
    def setup_custom_styles(self):
        """Setup custom styles for reports"""
        # Title style
        self.title_style = ParagraphStyle(
            'CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            alignment=TA_CENTER,
            textColor=colors.darkblue
        )
        
        # Heading style
        self.heading_style = ParagraphStyle(
            'CustomHeading',
            parent=self.styles['Heading2'],
            fontSize=16,
            spaceAfter=12,
            spaceBefore=20,
            textColor=colors.darkblue
        )
        
        # Subheading style
        self.subheading_style = ParagraphStyle(
            'CustomSubheading',
            parent=self.styles['Heading3'],
            fontSize=14,
            spaceAfter=8,
            spaceBefore=12,
            textColor=colors.darkgreen
        )
        
        # Normal text style
        self.normal_style = ParagraphStyle(
            'CustomNormal',
            parent=self.styles['Normal'],
            fontSize=11,
            spaceAfter=6
        )
        
        # Code style
        self.code_style = ParagraphStyle(
            'CustomCode',
            parent=self.styles['Code'],
            fontSize=9,
            fontName='Courier',
            leftIndent=20,
            rightIndent=20,
            backColor=colors.lightgrey
        )

    def generate_pdf_report(self, report_type="comprehensive", days=30, filters=None):
        """Generate a professional PDF report"""
        if not REPORTLAB_AVAILABLE:
            return None, "PDF generation requires reportlab library"
        
        try:
            # Get data
            indicators = self._get_filtered_data(days, filters)
            
            # Create filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"threat_intelligence_report_{report_type}_{timestamp}.pdf"
            filepath = os.path.join(self.reports_dir, filename)
            
            # Create PDF document
            doc = SimpleDocTemplate(filepath, pagesize=A4)
            story = []
            
            # Title
            story.append(Paragraph("Harmonia Incident Response", self.title_style))
            story.append(Paragraph(f"Threat Intelligence Report - {report_type.title()}", self.title_style))
            story.append(Spacer(1, 20))
            
            # Executive Summary
            story.append(Paragraph("Executive Summary", self.heading_style))
            story.append(Paragraph(
                f"This report provides a comprehensive analysis of threat intelligence data "
                f"collected over the past {days} days. The analysis covers {len(indicators)} "
                f"threat indicators from multiple sources including MITRE ATT&CK and CISA KEV catalog.",
                self.normal_style
            ))
            story.append(Spacer(1, 12))
            
            # Key Metrics
            story.append(Paragraph("Key Metrics", self.heading_style))
            metrics_data = self._calculate_metrics(indicators)
            metrics_table = self._create_metrics_table(metrics_data, self.styles)
            story.append(metrics_table)
            story.append(Spacer(1, 12))
            
            # Threat Analysis
            story.append(Paragraph("Threat Analysis", self.heading_style))
            threat_analysis = self._analyze_threats(indicators)
            story.append(Paragraph(threat_analysis, self.normal_style))
            story.append(Spacer(1, 12))
            
            # Top Threats Table
            story.append(Paragraph("Top Threats by Severity", self.heading_style))
            top_threats = sorted(indicators, key=lambda x: x.severity_score, reverse=True)[:10]
            threats_table = self._create_threats_table(top_threats, self.styles)
            story.append(threats_table)
            story.append(Spacer(1, 12))
            
            # Recommendations
            story.append(Paragraph("Recommendations", self.heading_style))
            recommendations = self._generate_recommendations(indicators)
            story.append(Paragraph(recommendations, self.normal_style))
            
            # Build PDF
            doc.build(story)
            
            return filename, None
            
        except Exception as e:
            return None, f"Error generating PDF report: {str(e)}"
    
    def generate_excel_report(self, report_type="comprehensive", days=30, filters=None):
        """Generate a professional Excel report"""
        if not OPENPYXL_AVAILABLE:
            return None, "Excel generation requires openpyxl library"
        
        try:
            # Get data
            indicators = self._get_filtered_data(days, filters)
            
            # Create filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"threat_intelligence_report_{report_type}_{timestamp}.xlsx"
            filepath = os.path.join(self.reports_dir, filename)
            
            # Create workbook
            wb = Workbook()
            
            # Summary sheet
            ws_summary = wb.active
            ws_summary.title = "Executive Summary"
            
            # Title
            ws_summary['A1'] = "Harmonia Incident Response"
            ws_summary['A2'] = f"Threat Intelligence Report - {report_type.title()}"
            ws_summary['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            
            # Style title
            title_font = Font(size=16, bold=True, color="1F4E79")
            ws_summary['A1'].font = title_font
            ws_summary['A2'].font = title_font
            
            # Key metrics
            ws_summary['A5'] = "Key Metrics"
            ws_summary['A5'].font = Font(bold=True)
            
            metrics_data = self._calculate_metrics(indicators)
            row = 6
            for metric, value in metrics_data.items():
                ws_summary[f'A{row}'] = metric
                ws_summary[f'B{row}'] = value
                row += 1
            
            # Detailed data sheet
            ws_data = wb.create_sheet("Threat Data")
            
            # Headers
            headers = ['ID', 'Name', 'Type', 'Description', 'Severity', 'Source', 'Date Added', 'Value']
            for col, header in enumerate(headers, 1):
                cell = ws_data.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
            
            # Data
            for row, indicator in enumerate(indicators, 2):
                ws_data.cell(row=row, column=1, value=indicator.id)
                ws_data.cell(row=row, column=2, value=indicator.name)
                ws_data.cell(row=row, column=3, value=indicator.indicator_type)
                ws_data.cell(row=row, column=4, value=indicator.description)
                ws_data.cell(row=row, column=5, value=indicator.severity_score)
                ws_data.cell(row=row, column=6, value=indicator.source)
                ws_data.cell(row=row, column=7, value=indicator.date_added)
                ws_data.cell(row=row, column=8, value=indicator.indicator_value)
            
            # Auto-adjust column widths
            for column in ws_data.columns:
                if column and len(column) > 0 and column[0] is not None:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        if cell is not None:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                    adjusted_width = min(max_length + 2, 50)
                    ws_data.column_dimensions[column_letter].width = adjusted_width
            
            # Analysis sheet
            ws_analysis = wb.create_sheet("Threat Analysis")
            ws_analysis['A1'] = "Threat Analysis"
            ws_analysis['A1'].font = Font(size=14, bold=True)
            
            analysis_text = self._analyze_threats(indicators)
            ws_analysis['A3'] = analysis_text
            
            # Save workbook
            wb.save(filepath)
            
            return filename, None
            
        except Exception as e:
            return None, f"Error generating Excel report: {str(e)}"
    
    def generate_html_report(self, report_type="comprehensive", days=30, filters=None):
        """Generate an HTML report"""
        try:
            # Get data
            indicators = self._get_filtered_data(days, filters)
            
            # Create filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"threat_intelligence_report_{report_type}_{timestamp}.html"
            filepath = os.path.join(self.reports_dir, filename)
            
            # Generate HTML content
            html_content = self._generate_html_content(indicators, report_type, days)
            
            # Save file
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            return filename, None
            
        except Exception as e:
            return None, f"Error generating HTML report: {str(e)}"
    
    def export_data(self, format_type="json", days=30, filters=None):
        """Export raw data in various formats"""
        try:
            # Get data
            indicators = self._get_filtered_data(days, filters)
            
            # Create filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            if format_type == "json":
                filename = f"threat_data_{timestamp}.json"
                filepath = os.path.join(self.reports_dir, filename)
                
                data = []
                for indicator in indicators:
                    data.append({
                        'id': indicator.id,
                        'name': indicator.name,
                        'type': indicator.indicator_type,
                        'description': indicator.description,
                        'severity_score': indicator.severity_score,
                        'source': indicator.source,
                        'date_added': indicator.date_added,
                        'indicator_value': indicator.indicator_value
                    })
                
                with open(filepath, 'w', encoding='utf-8') as f:
                    json.dump(data, f, indent=2, default=str)
            
            elif format_type == "csv":
                filename = f"threat_data_{timestamp}.csv"
                filepath = os.path.join(self.reports_dir, filename)
                
                with open(filepath, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    writer.writerow(['ID', 'Name', 'Type', 'Description', 'Severity', 'Source', 'Date Added', 'Value'])
                    
                    for indicator in indicators:
                        writer.writerow([
                            indicator.id,
                            indicator.name,
                            indicator.indicator_type,
                            indicator.description,
                            indicator.severity_score,
                            indicator.source,
                            indicator.date_added,
                            indicator.indicator_value
                        ])
            
            return filename, None
            
        except Exception as e:
            return None, f"Error exporting data: {str(e)}"
    
    def _get_filtered_data(self, days=30, filters=None):
        """Get filtered indicator data"""
        query = Indicator.query
        
        # Apply date filter
        if days:
            cutoff_date = datetime.now() - timedelta(days=days)
            query = query.filter(Indicator.date_added >= cutoff_date.strftime('%Y-%m-%d'))
        
        # Apply additional filters
        if filters:
            if filters.get('type'):
                query = query.filter(Indicator.indicator_type == filters['type'])
            if filters.get('source'):
                query = query.filter(Indicator.source == filters['source'])
            if filters.get('severity_min'):
                query = query.filter(Indicator.severity_score >= filters['severity_min'])
            if filters.get('severity_max'):
                query = query.filter(Indicator.severity_score <= filters['severity_max'])
        
        return query.all()
    
    def _calculate_metrics(self, indicators):
        """Calculate key metrics from indicators"""
        if not indicators:
            return {}
        
        total_indicators = len(indicators)
        
        # Convert severity scores to float, handling both string and numeric values
        severity_scores = []
        for ind in indicators:
            try:
                score = float(ind.severity_score) if ind.severity_score else 0.0
                severity_scores.append(score)
            except (ValueError, TypeError):
                severity_scores.append(0.0)
        
        avg_severity = sum(severity_scores) / total_indicators if severity_scores else 0.0
        
        # Count by type
        type_counts = {}
        for ind in indicators:
            type_counts[ind.indicator_type] = type_counts.get(ind.indicator_type, 0) + 1
        
        # Count by source
        source_counts = {}
        for ind in indicators:
            source_counts[ind.source] = source_counts.get(ind.source, 0) + 1
        
        # High severity indicators (severity >= 7)
        high_severity = len([score for score in severity_scores if score >= 7])
        
        return {
            "Total Indicators": total_indicators,
            "Average Severity": f"{avg_severity:.2f}",
            "High Severity Indicators": high_severity,
            "Most Common Type": max(type_counts.items(), key=lambda x: x[1])[0] if type_counts else "N/A",
            "Most Active Source": max(source_counts.items(), key=lambda x: x[1])[0] if source_counts else "N/A"
        }
    
    def _create_metrics_table(self, metrics_data, styles):
        """Create a table for metrics data"""
        data = [[Paragraph("Metric", styles['Heading3']), Paragraph("Value", styles['Heading3'])]]
        for metric, value in metrics_data.items():
            data.append([Paragraph(metric, styles['Normal']), Paragraph(str(value), styles['Normal'])])
        
        table = Table(data, colWidths=[3*inch, 2*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        return table
    
    def _create_threats_table(self, threats, styles):
        """Create a table for threats data"""
        data = [[
            Paragraph("Name", styles['Heading3']),
            Paragraph("Type", styles['Heading3']),
            Paragraph("Severity", styles['Heading3']),
            Paragraph("Source", styles['Heading3'])
        ]]
        
        for threat in threats:
            data.append([
                Paragraph(threat.name[:50] + "..." if len(threat.name) > 50 else threat.name, styles['Normal']),
                Paragraph(threat.indicator_type, styles['Normal']),
                Paragraph(str(threat.severity_score), styles['Normal']),
                Paragraph(threat.source, styles['Normal'])
            ])
        
        table = Table(data, colWidths=[2.5*inch, 1.5*inch, 1*inch, 1.5*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        return table
    
    def _analyze_threats(self, indicators):
        """Generate threat analysis text"""
        if not indicators:
            return "No threat data available for analysis."
        
        # Calculate statistics
        total = len(indicators)
        
        # Convert severity scores to float, handling both string and numeric values
        severity_scores = []
        for ind in indicators:
            try:
                score = float(ind.severity_score) if ind.severity_score else 0.0
                severity_scores.append(score)
            except (ValueError, TypeError):
                severity_scores.append(0.0)
        
        avg_severity = sum(severity_scores) / total if severity_scores else 0.0
        high_severity = len([score for score in severity_scores if score >= 7])
        
        # Most common types
        type_counts = {}
        for ind in indicators:
            type_counts[ind.indicator_type] = type_counts.get(ind.indicator_type, 0) + 1
        
        most_common_type = max(type_counts.items(), key=lambda x: x[1])
        
        analysis = f"""
        Threat Analysis Summary:
        
        Total Indicators Analyzed: {total}
        Average Severity Score: {avg_severity:.2f}/10
        High Severity Threats (≥7): {high_severity} ({high_severity/total*100:.1f}%)
        Most Common Threat Type: {most_common_type[0]} ({most_common_type[1]} instances)
        
        Key Findings:
        • {high_severity} high-severity threats require immediate attention
        • Average threat severity of {avg_severity:.2f} indicates moderate overall risk
        • {most_common_type[0]} threats are the most prevalent, suggesting focused defense needed
        
        Recommendations:
        • Prioritize response to high-severity threats
        • Implement specific defenses against {most_common_type[0]} threats
        • Consider threat hunting for related indicators
        • Review and update security controls based on threat patterns
        """
        
        return analysis
    
    def _generate_recommendations(self, indicators):
        """Generate security recommendations"""
        if not indicators:
            return "No recommendations available due to insufficient data."
        
        recommendations = """
        Security Recommendations:
        
        1. Immediate Actions:
           • Review and respond to all high-severity threats
           • Update threat detection rules based on observed patterns
           • Conduct threat hunting for related indicators
        
        2. Strategic Improvements:
           • Enhance monitoring for most common threat types
           • Implement additional security controls where gaps exist
           • Develop incident response playbooks for observed threats
        
        3. Long-term Planning:
           • Regular threat intelligence updates and analysis
           • Continuous improvement of security posture
           • Staff training on emerging threats and response procedures
        """
        
        return recommendations
    
    def _generate_html_content(self, indicators, report_type, days):
        """Generate HTML content for reports"""
        metrics_data = self._calculate_metrics(indicators)
        
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Harmonia Threat Intelligence Report</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 40px; }}
                .header {{ text-align: center; color: #1F4E79; margin-bottom: 30px; }}
                .section {{ margin: 20px 0; }}
                .metrics {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 20px; margin: 20px 0; }}
                .metric {{ background: #f5f5f5; padding: 15px; border-radius: 5px; text-align: center; }}
                .metric-value {{ font-size: 24px; font-weight: bold; color: #1F4E79; }}
                table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                th {{ background-color: #1F4E79; color: white; }}
                tr:nth-child(even) {{ background-color: #f2f2f2; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>Harmonia Incident Response</h1>
                <h2>Threat Intelligence Report - {report_type.title()}</h2>
                <p>Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
            </div>
            
            <div class="section">
                <h3>Executive Summary</h3>
                <p>This report provides a comprehensive analysis of threat intelligence data collected over the past {days} days. 
                The analysis covers {len(indicators)} threat indicators from multiple sources including MITRE ATT&CK and CISA KEV catalog.</p>
            </div>
            
            <div class="section">
                <h3>Key Metrics</h3>
                <div class="metrics">
        """
        
        for metric, value in metrics_data.items():
            html += f"""
                    <div class="metric">
                        <div class="metric-value">{value}</div>
                        <div>{metric}</div>
                    </div>
            """
        
        html += """
                </div>
            </div>
            
            <div class="section">
                <h3>Top Threats by Severity</h3>
                <table>
                    <tr>
                        <th>Name</th>
                        <th>Type</th>
                        <th>Severity</th>
                        <th>Source</th>
                    </tr>
        """
        
        top_threats = sorted(indicators, key=lambda x: x.severity_score, reverse=True)[:10]
        for threat in top_threats:
            html += f"""
                    <tr>
                        <td>{threat.name}</td>
                        <td>{threat.indicator_type}</td>
                        <td>{threat.severity_score}</td>
                        <td>{threat.source}</td>
                    </tr>
            """
        
        html += """
                </table>
            </div>
            
            <div class="section">
                <h3>Threat Analysis</h3>
                <p>
        """
        
        html += self._analyze_threats(indicators).replace('\n', '<br>')
        
        html += """
                </p>
            </div>
            
            <div class="section">
                <h3>Recommendations</h3>
                <p>
        """
        
        html += self._generate_recommendations(indicators).replace('\n', '<br>')
        
        html += """
                </p>
            </div>
        </body>
        </html>
        """
        
        return html 